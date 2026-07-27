"""
tests/conftest.py
Root-level shared fixtures and hooks for the Blogify E2E test suite.
Applies to both Selenium (web) and Appium (mobile) test runs.
"""
import os
import json
import time
import pytest


# ── Report directory bootstrap ──────────────────────────────────────────────

def pytest_configure(config):
    """Create the reports/ directory before any test is collected."""
    reports_dir = os.path.join(os.path.dirname(__file__), "reports")
    os.makedirs(reports_dir, exist_ok=True)


# ── Session-level results accumulator ──────────────────────────────────────

_results: list[dict] = []


def pytest_runtest_logreport(report):
    """Collect per-test outcomes for the final Excel/JSON report."""
    if report.when == "call":
        _results.append({
            "nodeid":    report.nodeid,
            "outcome":   report.outcome,          # "passed" | "failed" | "error"
            "duration":  round(report.duration, 3),
            "longrepr":  str(report.longrepr) if report.failed else "",
        })


# ── Session-end: write JSON summary ────────────────────────────────────────

def pytest_sessionfinish(session, exitstatus):
    """Write a JSON summary and attempt to generate an Excel report."""
    reports_dir = os.path.join(os.path.dirname(__file__), "reports")
    json_path = os.path.join(reports_dir, "test_results.json")

    summary = {
        "generated_at": time.strftime("%Y-%m-%d %H:%M:%S"),
        "total":  len(_results),
        "passed": sum(1 for r in _results if r["outcome"] == "passed"),
        "failed": sum(1 for r in _results if r["outcome"] == "failed"),
        "error":  sum(1 for r in _results if r["outcome"] == "error"),
        "results": _results,
    }

    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(summary, f, indent=2)

    # Attempt Excel generation (requires openpyxl)
    try:
        _generate_excel_report(summary, reports_dir)
    except Exception as exc:
        print(f"\n[report] Excel generation skipped: {exc}")


def _generate_excel_report(summary: dict, reports_dir: str):
    """Generate an Excel workbook from the accumulated test results."""
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # ── Summary sheet ───────────────────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Summary"

    header_fill   = PatternFill("solid", fgColor="6D28D9")   # violet-700
    pass_fill     = PatternFill("solid", fgColor="16A34A")   # green-600
    fail_fill     = PatternFill("solid", fgColor="DC2626")   # red-600
    error_fill    = PatternFill("solid", fgColor="D97706")   # amber-600
    white_font    = Font(color="FFFFFF", bold=True)
    thin_border   = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Merge title
    ws_summary.merge_cells("A1:E1")
    title_cell = ws_summary["A1"]
    title_cell.value = "Blogify — E2E Test Report"
    title_cell.font  = Font(bold=True, size=16, color="FFFFFF")
    title_cell.fill  = header_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_summary.row_dimensions[1].height = 30

    # Meta info
    meta = [
        ("Generated At", summary["generated_at"]),
        ("Total Tests",  summary["total"]),
        ("Passed",       summary["passed"]),
        ("Failed",       summary["failed"]),
        ("Errors",       summary["error"]),
        ("Pass Rate",    f"{summary['passed'] / max(summary['total'], 1) * 100:.1f}%"),
    ]
    for i, (label, value) in enumerate(meta, start=3):
        ws_summary[f"A{i}"] = label
        ws_summary[f"A{i}"].font = Font(bold=True)
        ws_summary[f"B{i}"] = value

    # ── Results sheet ───────────────────────────────────────────────────────
    ws_results = wb.create_sheet(title="All Results")

    headers = ["#", "Test ID", "Node ID", "Outcome", "Duration (s)", "Failure Detail"]
    for col, h in enumerate(headers, start=1):
        cell = ws_results.cell(row=1, column=col, value=h)
        cell.font  = white_font
        cell.fill  = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for row_idx, result in enumerate(summary["results"], start=2):
        # Derive a short TC-ID from the test function name if possible
        parts = result["nodeid"].split("::")
        tc_id = parts[-1].split("_")[0:3]
        tc_id_str = "_".join(tc_id).upper() if len(tc_id) >= 2 else parts[-1]

        outcome  = result["outcome"].upper()
        fill = pass_fill if outcome == "PASSED" else (fail_fill if outcome == "FAILED" else error_fill)

        row_data = [row_idx - 1, tc_id_str, result["nodeid"], outcome, result["duration"], result["longrepr"]]
        for col, val in enumerate(row_data, start=1):
            cell = ws_results.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if col == 4:
                cell.fill = fill
                cell.font = Font(bold=True, color="FFFFFF")

    # Column widths
    col_widths = [5, 25, 60, 12, 15, 60]
    for i, width in enumerate(col_widths, start=1):
        ws_results.column_dimensions[get_column_letter(i)].width = width

    # Freeze header row
    ws_results.freeze_panes = "A2"

    excel_path = os.path.join(reports_dir, "Blogify_E2E_Test_Report.xlsx")
    wb.save(excel_path)
    print(f"\n[report] Excel report saved -> {excel_path}")
