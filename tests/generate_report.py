#!/usr/bin/env python3
"""
tests/generate_report.py
──────────────────────────────────────────────────────────────────────────────
Standalone script to generate an Excel + HTML summary from a previously
produced test_results.json.

Usage:
    python generate_report.py                            # reads reports/test_results.json
    python generate_report.py --json path/to/file.json  # custom JSON path
──────────────────────────────────────────────────────────────────────────────
"""
import argparse
import json
import os
import sys
import time

# ── Resolve paths ────────────────────────────────────────────────────────────

TESTS_DIR   = os.path.dirname(os.path.abspath(__file__))
REPORTS_DIR = os.path.join(TESTS_DIR, "reports")


def load_results(json_path: str) -> dict:
    with open(json_path, "r", encoding="utf-8") as f:
        return json.load(f)


# ── Excel generation ─────────────────────────────────────────────────────────

def generate_excel(summary: dict, out_path: str):
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.chart import BarChart, Reference
    except ImportError:
        print("[ERROR] openpyxl not installed. Run: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.Workbook()

    # ─ Colours / styles ──────────────────────────────────────────────────────
    violet_fill  = PatternFill("solid", fgColor="6D28D9")
    green_fill   = PatternFill("solid", fgColor="16A34A")
    red_fill     = PatternFill("solid", fgColor="DC2626")
    amber_fill   = PatternFill("solid", fgColor="D97706")
    sky_fill     = PatternFill("solid", fgColor="0284C7")
    white_font   = Font(color="FFFFFF", bold=True, size=11)
    header_font  = Font(color="FFFFFF", bold=True, size=10)
    thin_border  = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center")

    results = summary.get("results", [])
    total   = summary.get("total", len(results))
    passed  = summary.get("passed", sum(1 for r in results if r["outcome"] == "passed"))
    failed  = summary.get("failed", sum(1 for r in results if r["outcome"] == "failed"))
    error   = summary.get("error",  sum(1 for r in results if r["outcome"] == "error"))
    pass_rate = f"{passed / max(total, 1) * 100:.1f}%"

    # ─────────────────────────────────────────────────────────────────────────
    # Sheet 1 — Dashboard / Summary
    # ─────────────────────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_view.showGridLines = False

    # Big title
    ws.merge_cells("A1:G1")
    t = ws["A1"]
    t.value     = "🚀  Blogify — E2E Test Report"
    t.font      = Font(bold=True, size=18, color="FFFFFF")
    t.fill      = violet_fill
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 42

    # Sub-title
    ws.merge_cells("A2:G2")
    s = ws["A2"]
    s.value     = f"Generated: {summary.get('generated_at', time.strftime('%Y-%m-%d %H:%M:%S'))}"
    s.font      = Font(italic=True, size=10, color="6D28D9")
    s.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 20

    # Stat cards (row 4–5)
    stats = [
        ("Total Tests", total,     sky_fill),
        ("Passed",      passed,    green_fill),
        ("Failed",      failed,    red_fill),
        ("Errors",      error,     amber_fill),
        ("Pass Rate",   pass_rate, violet_fill),
    ]
    for col_offset, (label, value, fill) in enumerate(stats):
        col = col_offset * 1 + 1          # columns A,B,C,D,E
        label_cell = ws.cell(row=4, column=col, value=label)
        label_cell.font      = Font(bold=True, size=9, color="FFFFFF")
        label_cell.fill      = fill
        label_cell.alignment = center_align
        label_cell.border    = thin_border
        ws.row_dimensions[4].height = 18

        value_cell = ws.cell(row=5, column=col, value=value)
        value_cell.font      = Font(bold=True, size=14, color="FFFFFF")
        value_cell.fill      = fill
        value_cell.alignment = center_align
        value_cell.border    = thin_border
        ws.row_dimensions[5].height = 28

    # Column widths
    for c in range(1, 8):
        ws.column_dimensions[get_column_letter(c)].width = 18

    # ─────────────────────────────────────────────────────────────────────────
    # Sheet 2 — All Results
    # ─────────────────────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("All Results")
    ws2.sheet_view.showGridLines = False

    headers = ["#", "TC Reference", "Full Node ID", "Suite", "Outcome", "Duration (s)", "Failure Detail"]
    for col, h in enumerate(headers, start=1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font      = header_font
        cell.fill      = violet_fill
        cell.alignment = center_align
        cell.border    = thin_border
    ws2.row_dimensions[1].height = 20
    ws2.freeze_panes = "A2"

    for row_idx, result in enumerate(results, start=2):
        nodeid   = result.get("nodeid", "")
        outcome  = result.get("outcome", "").upper()
        duration = result.get("duration", 0.0)
        longrepr = result.get("longrepr", "")

        # Infer short TC id from function name
        parts = nodeid.split("::")
        fn    = parts[-1] if parts else nodeid
        tokens = fn.split("_")
        tc_ref = "_".join(tokens[:4]).upper()

        # Detect suite (selenium / appium)
        if "selenium" in nodeid.lower():
            suite = "Selenium (Web)"
        elif "appium" in nodeid.lower():
            suite = "Appium (Mobile)"
        else:
            suite = "Unknown"

        fill_map = {"PASSED": green_fill, "FAILED": red_fill, "ERROR": amber_fill}
        row_fill = fill_map.get(outcome, PatternFill())

        row_data = [row_idx - 1, tc_ref, nodeid, suite, outcome, duration, longrepr]
        for col, val in enumerate(row_data, start=1):
            cell = ws2.cell(row=row_idx, column=col, value=val)
            cell.border    = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if col == 5:
                cell.fill = row_fill
                cell.font = Font(bold=True, color="FFFFFF", size=9)

    # Column widths
    col_widths = [4, 28, 65, 18, 12, 14, 55]
    for i, w in enumerate(col_widths, start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ─────────────────────────────────────────────────────────────────────────
    # Sheet 3 — Failed Tests
    # ─────────────────────────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Failed Tests")
    ws3.sheet_view.showGridLines = False
    failed_results = [r for r in results if r.get("outcome") != "passed"]

    for col, h in enumerate(["#", "TC Reference", "Node ID", "Outcome", "Duration (s)", "Failure Detail"], start=1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font      = header_font
        cell.fill      = red_fill
        cell.alignment = center_align
        cell.border    = thin_border
    ws3.freeze_panes = "A2"

    if not failed_results:
        ws3.merge_cells("A2:F2")
        c = ws3["A2"]
        c.value     = "✅  All tests passed — no failures to report!"
        c.font      = Font(bold=True, color="16A34A", size=12)
        c.alignment = Alignment(horizontal="center")
    else:
        for row_idx, result in enumerate(failed_results, start=2):
            nodeid   = result.get("nodeid", "")
            outcome  = result.get("outcome", "").upper()
            duration = result.get("duration", 0.0)
            longrepr = result.get("longrepr", "")
            parts    = nodeid.split("::")
            fn       = parts[-1] if parts else nodeid
            tc_ref   = "_".join(fn.split("_")[:4]).upper()
            fill     = red_fill if outcome == "FAILED" else amber_fill

            for col, val in enumerate([row_idx - 1, tc_ref, nodeid, outcome, duration, longrepr], start=1):
                cell = ws3.cell(row=row_idx, column=col, value=val)
                cell.border    = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                if col == 4:
                    cell.fill = fill
                    cell.font = Font(bold=True, color="FFFFFF", size=9)

    for i, w in enumerate([4, 28, 65, 12, 14, 55], start=1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    wb.save(out_path)
    print(f"[SUCCESS] Excel report saved -> {out_path}")


# ── HTML generation ───────────────────────────────────────────────────────────

def generate_html(summary: dict, out_path: str):
    results  = summary.get("results", [])
    total    = summary.get("total",  len(results))
    passed   = summary.get("passed", sum(1 for r in results if r["outcome"] == "passed"))
    failed   = summary.get("failed", sum(1 for r in results if r["outcome"] == "failed"))
    error    = summary.get("error",  sum(1 for r in results if r["outcome"] == "error"))
    gen_at   = summary.get("generated_at", time.strftime("%Y-%m-%d %H:%M:%S"))
    pass_pct = round(passed / max(total, 1) * 100, 1)

    rows_html = ""
    for i, r in enumerate(results, start=1):
        outcome  = r.get("outcome", "")
        cls      = {"passed": "pass", "failed": "fail", "error": "err"}.get(outcome, "")
        label    = outcome.upper()
        nodeid   = r.get("nodeid", "")
        duration = r.get("duration", 0.0)
        detail   = (r.get("longrepr") or "—").replace("<", "&lt;").replace(">", "&gt;").replace("\n", "<br>")
        suite    = "Selenium" if "selenium" in nodeid.lower() else ("Appium" if "appium" in nodeid.lower() else "—")
        fn       = nodeid.split("::")[-1] if "::" in nodeid else nodeid
        tc_ref   = "_".join(fn.split("_")[:4]).upper()
        rows_html += f"""
        <tr class="{cls}">
            <td>{i}</td>
            <td title="{nodeid}">{tc_ref}</td>
            <td>{suite}</td>
            <td><span class="badge {cls}">{label}</span></td>
            <td>{duration:.3f}s</td>
            <td class="detail">{detail}</td>
        </tr>"""

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Blogify — E2E Test Report</title>
<style>
  :root{{--violet:#6d28d9;--green:#16a34a;--red:#dc2626;--amber:#d97706;--sky:#0284c7}}
  *{{box-sizing:border-box;margin:0;padding:0}}
  body{{font-family:'Segoe UI',system-ui,sans-serif;background:#f1f5f9;color:#1e293b}}
  header{{background:var(--violet);color:#fff;padding:28px 40px}}
  header h1{{font-size:1.8rem;font-weight:800;letter-spacing:-.5px}}
  header p{{font-size:.85rem;opacity:.75;margin-top:4px}}
  .cards{{display:flex;gap:16px;padding:24px 40px;flex-wrap:wrap}}
  .card{{flex:1;min-width:140px;background:#fff;border-radius:12px;padding:20px;
         box-shadow:0 1px 4px rgba(0,0,0,.06);text-align:center}}
  .card .num{{font-size:2rem;font-weight:800;line-height:1}}
  .card .lbl{{font-size:.75rem;color:#64748b;margin-top:6px;text-transform:uppercase;letter-spacing:.05em}}
  .card.t{{border-top:4px solid var(--sky)}}  .card.t .num{{color:var(--sky)}}
  .card.p{{border-top:4px solid var(--green)}} .card.p .num{{color:var(--green)}}
  .card.f{{border-top:4px solid var(--red)}}  .card.f .num{{color:var(--red)}}
  .card.e{{border-top:4px solid var(--amber)}} .card.e .num{{color:var(--amber)}}
  .card.r{{border-top:4px solid var(--violet)}} .card.r .num{{color:var(--violet)}}
  .progress-wrap{{padding:0 40px 24px}}
  .progress-bar{{height:10px;background:#e2e8f0;border-radius:99px;overflow:hidden}}
  .progress-fill{{height:100%;background:var(--green);border-radius:99px;transition:width .6s ease}}
  .progress-label{{font-size:.8rem;color:#64748b;margin-top:6px}}
  section{{padding:0 40px 40px}}
  .toolbar{{display:flex;gap:8px;margin-bottom:12px;flex-wrap:wrap;align-items:center}}
  .toolbar input{{padding:7px 12px;border:1px solid #e2e8f0;border-radius:8px;font-size:.85rem;flex:1;min-width:200px}}
  .filter-btn{{padding:6px 14px;border:1px solid #e2e8f0;border-radius:8px;cursor:pointer;font-size:.82rem;background:#fff;transition:.15s}}
  .filter-btn.active,.filter-btn:hover{{background:var(--violet);color:#fff;border-color:var(--violet)}}
  table{{width:100%;border-collapse:collapse;background:#fff;border-radius:12px;overflow:hidden;box-shadow:0 1px 4px rgba(0,0,0,.06)}}
  th{{background:var(--violet);color:#fff;text-align:left;padding:12px 14px;font-size:.8rem;text-transform:uppercase;letter-spacing:.05em}}
  td{{padding:10px 14px;font-size:.82rem;border-bottom:1px solid #f1f5f9;vertical-align:top}}
  tr:last-child td{{border-bottom:none}}
  tr.pass td:first-child{{border-left:4px solid var(--green)}}
  tr.fail td:first-child{{border-left:4px solid var(--red)}}
  tr.err  td:first-child{{border-left:4px solid var(--amber)}}
  tr:hover td{{background:#f8fafc}}
  .badge{{display:inline-block;padding:2px 10px;border-radius:999px;font-size:.72rem;font-weight:700;text-transform:uppercase;letter-spacing:.04em}}
  .badge.pass{{background:#dcfce7;color:var(--green)}}
  .badge.fail{{background:#fee2e2;color:var(--red)}}
  .badge.err {{background:#fef3c7;color:var(--amber)}}
  .detail{{font-family:monospace;font-size:.75rem;color:#64748b;max-width:400px;word-break:break-word}}
  footer{{text-align:center;padding:20px;font-size:.78rem;color:#94a3b8}}
</style>
</head>
<body>
<header>
  <h1>🚀 Blogify — E2E Test Report</h1>
  <p>Generated: {gen_at}</p>
</header>

<div class="cards">
  <div class="card t"><div class="num">{total}</div><div class="lbl">Total Tests</div></div>
  <div class="card p"><div class="num">{passed}</div><div class="lbl">Passed</div></div>
  <div class="card f"><div class="num">{failed}</div><div class="lbl">Failed</div></div>
  <div class="card e"><div class="num">{error}</div><div class="lbl">Errors</div></div>
  <div class="card r"><div class="num">{pass_pct}%</div><div class="lbl">Pass Rate</div></div>
</div>

<div class="progress-wrap">
  <div class="progress-bar"><div class="progress-fill" style="width:{pass_pct}%"></div></div>
  <div class="progress-label">{passed} / {total} tests passed</div>
</div>

<section>
  <div class="toolbar">
    <input type="text" id="search" placeholder="🔍 Filter by TC reference or node ID…" oninput="filterTable()">
    <button class="filter-btn active" onclick="filterSuite('all',this)">All</button>
    <button class="filter-btn" onclick="filterSuite('Selenium',this)">Selenium</button>
    <button class="filter-btn" onclick="filterSuite('Appium',this)">Appium</button>
    <button class="filter-btn" onclick="filterOutcome('pass',this)">✅ Passed</button>
    <button class="filter-btn" onclick="filterOutcome('fail',this)">❌ Failed</button>
    <button class="filter-btn" onclick="filterOutcome('err',this)">⚠️ Errors</button>
  </div>
  <table id="results-table">
    <thead>
      <tr><th>#</th><th>TC Reference</th><th>Suite</th><th>Outcome</th><th>Duration</th><th>Failure Detail</th></tr>
    </thead>
    <tbody>
      {rows_html}
    </tbody>
  </table>
</section>

<footer>Blogify E2E Testing Suite · {total} tests · Generated {gen_at}</footer>

<script>
  let activeSuite = 'all', activeOutcome = '';

  function filterTable() {{
    const query = document.getElementById('search').value.toLowerCase();
    document.querySelectorAll('#results-table tbody tr').forEach(tr => {{
      const text = tr.innerText.toLowerCase();
      const suiteCell = tr.cells[2]?.innerText || '';
      const cls = tr.className;
      const suiteMatch = activeSuite === 'all' || suiteCell.includes(activeSuite);
      const outcomeMatch = !activeOutcome || cls === activeOutcome;
      tr.style.display = text.includes(query) && suiteMatch && outcomeMatch ? '' : 'none';
    }});
  }}

  function filterSuite(suite, btn) {{
    activeSuite = suite;
    document.querySelectorAll('.filter-btn').forEach(b => b.classList.remove('active'));
    btn.classList.add('active');
    filterTable();
  }}

  function filterOutcome(outcome, btn) {{
    if (activeOutcome === outcome) {{ activeOutcome = ''; btn.classList.remove('active'); }}
    else {{
      activeOutcome = outcome;
      document.querySelectorAll('.filter-btn').forEach(b => {{
        if (['pass','fail','err'].some(o => b.getAttribute('onclick')?.includes(o))) b.classList.remove('active');
      }});
      btn.classList.add('active');
    }}
    filterTable();
  }}
</script>
</body>
</html>"""

    with open(out_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"[SUCCESS] HTML  report saved -> {out_path}")


# ── Entry point ───────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Generate Blogify test reports from JSON results.")
    parser.add_argument("--json", default=os.path.join(REPORTS_DIR, "test_results.json"),
                        help="Path to test_results.json (default: reports/test_results.json)")
    parser.add_argument("--excel", default=os.path.join(REPORTS_DIR, "Blogify_E2E_Test_Report.xlsx"),
                        help="Output path for the Excel report")
    parser.add_argument("--html", default=os.path.join(REPORTS_DIR, "blogify_html_report.html"),
                        help="Output path for the HTML report")
    args = parser.parse_args()

    if not os.path.exists(args.json):
        print(f"[ERROR] JSON results file not found: {args.json}")
        print("        Run pytest first to generate test_results.json")
        sys.exit(1)

    os.makedirs(REPORTS_DIR, exist_ok=True)
    summary = load_results(args.json)
    generate_excel(summary, args.excel)
    generate_html(summary, args.html)
    print("\n[SUCCESS] Report generation complete!")


if __name__ == "__main__":
    main()
